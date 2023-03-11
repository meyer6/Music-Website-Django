from unicodedata import name
from django.shortcuts import render
from django.http import HttpResponse
import account.views as views2
from account.forms import UsersForm
from account.models import Users
from django.http import HttpResponseRedirect
import re
from datetime import datetime as time
import openpyxl 
import copy

def songs_create():
    workbook = openpyxl.load_workbook(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Songs.xlsx")
    sheet = workbook.active
    songs = []
    num = "1"
    while sheet["A"+num].value!=None:
        songs.append([sheet["A"+num].value,sheet["B"+num].value,sheet["C"+num].value,sheet["D"+num].value,sheet["E"+num].value,sheet["F"+num].value])
        num = str(int(num)+1)
    return divide(songs)
def divide(songs):
    r=[0,0,0]
    for i in range(3):
        if len(songs)%4 >= i+1:
            for a in range(i,len(r)):
                r[a] = r[a] + 1
    songs1 = songs[:len(songs)//4+r[0]]
    songs2 = songs[len(songs)//4+r[0]:len(songs)//4*2+r[1]]
    songs3 = songs[len(songs)//4*2+r[1]:len(songs)//4*3+r[2]]
    songs4 = songs[len(songs)//4*3+r[2]:]
    return songs1,songs2,songs3,songs4
def Valid_details(password,date):
    if len(password)<8 :
        return "Password must be more than 8 characters"
    if bool(re.match(r'\w*[A-Z]\w*', password)) == False:
        return "Password must have upper case letters"
    if bool(re.match(r'\w*[a-z]\w*', password)) == False:
        return "Password must have lower case letters"        
    if any(char.isdigit() for char in password) == False:
        return "Password must have numbers"
    if int(date.replace("-","")) >= int(str(time.today().strftime('%Y-%m-%d')).replace("-","")):
        return "Please input valid date"
    return True

songs1,songs2,songs3,songs4 = songs_create()
all_songs = songs1 + songs2 + songs3 + songs4
playlist_songs = []
playlist = []
playing_songs = []
playing_songs_num = -1
name = ""

def profile(request):
    user1 = views2.user1
    issues = ""
    if user1 == []:
        return HttpResponseRedirect('/account')
    if request.method == 'POST':  # data sent by user
        form = UsersForm(request.POST)
        if form.is_valid():
            if Valid_details(request.POST.get("password", ""), request.POST.get("date", "")) == True: 
                record = Users.objects.get(email = user1[0])
                record.delete()
                form.save()
                user1 = [request.POST.get("email", ""), request.POST.get("password", ""), request.POST.get("date", ""), request.POST.get("fav_artist", ""), request.POST.get("fav_genre", "")]
            else:
                issues = Valid_details(request.POST.get("password", ""), request.POST.get("date", ""))
    return render(request, "Profile.html", { 'issue': issues, 'email': user1[0], 'password': user1[1], 'date': user1[2], 'fav_artist': user1[3], 'fav_genre': user1[4]})
def main(request):
    global playing_songs
    global playing_songs_num

    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
       
    if request.method == 'POST':
        songs1,songs2,songs3,songs4 = songs_create()
        songs = songs1 + songs2 + songs3 + songs4

        songs_backup = copy.deepcopy(songs)
        for i in range(len(songs_backup)-1,-1,-1):
            if request.POST.get("search", "").lower() not in songs_backup[i][0].lower() and request.POST.get("search", "").lower() not in songs_backup[i][1].lower():
                songs_backup.pop(i)
        songs1,songs2,songs3,songs4 = divide(songs_backup)
    else:
        songs1,songs2,songs3,songs4 = songs_create()
    playing_songs = songs1 + songs2 + songs3 + songs4

    file = open(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\static\Songs.txt","r+")
    file. truncate(0)
    file. close()
    myFile = open(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\static\Songs.txt","a")
    for song in playing_songs:
        for thing in song:
            myFile.write(thing+",")
        myFile.write("\n")
    myFile. close()
    return render(request, "Main.html", { 'songs1': songs1, 'songs2': songs2, 'songs3': songs3, 'songs4': songs4})
def organise(songs):
    songs1 = []
    songs2 = []
    songs3 = []
    songs4 = []
    a=0 
    for i in range(len(playing_songs)):
        if a%4 == 0:
            songs1.append(playing_songs[i])
        if a%4 == 1:
            songs2.append(playing_songs[i])
        if a%4 == 2:
            songs3.append(playing_songs[i])
        if a%4 == 3:
            songs4.append(playing_songs[i])
        a=a+1
    return songs1, songs2, songs3, songs4
def sort(request):
    global playing_songs
    global playing_songs_num

    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
    if "A-Z" == request.POST.get("sort", ""):
        sorted=False
        while sorted==False:
            sorted=True
            for i in range(len(playing_songs)-1):
                if playing_songs[i][0]>playing_songs[i+1][0]:
                    playing_songs[i],playing_songs[i+1]=playing_songs[i+1],playing_songs[i]
                    sorted=False     
        songs1, songs2, songs3, songs4 = organise(playing_songs)
    elif "Z-A" == request.POST.get("sort", ""):
        sorted=False
        while sorted==False:
            sorted=True
            for i in range(len(playing_songs)-1):
                if playing_songs[i][0]<playing_songs[i+1][0]:
                    playing_songs[i],playing_songs[i+1]=playing_songs[i+1],playing_songs[i]
                    sorted=False     
        songs1, songs2, songs3, songs4 = organise(playing_songs)
    else:
        songs1, songs2, songs3, songs4 = songs_create()
    return render(request, "Main.html", { 'songs1': songs1, 'songs2': songs2, 'songs3': songs3, 'songs4': songs4})
def save(request):
    return HttpResponseRedirect('/main')

def play_songs(request):
    global playing_songs
    global playing_songs_num
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
    if request.method == 'POST':
        if "next" in request.POST:
            playing_songs_num = (playing_songs_num + 1) % len(playing_songs)
        elif "previous" in request.POST:
            playing_songs_num = (playing_songs_num - 1 + len(playing_songs)) % len(playing_songs)
        else:
            workbook = openpyxl.load_workbook(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Songs.xlsx")
            sheet = workbook.active
            for i in range(len(playing_songs)):
                if playing_songs[i][0] in request.POST:
                    playing_songs_num = i
                    break
    return render(request, "Play_songs.html", {'name' : playing_songs[playing_songs_num][0], 'artist' : playing_songs[playing_songs_num][1], 'song' : playing_songs[playing_songs_num][2], 'image' : playing_songs[playing_songs_num][3]})

def playlists(request):
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
    global playlist_songs
    playlist_songs = []
    workbook = openpyxl.load_workbook(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Playlists.xlsx")
    sheet = workbook.active
    playlists = []
    num = "1"
    while sheet["A"+num].value!=None:
        if sheet["A"+num].value == user1[0]:
            playlists.append([sheet["A"+num].value, sheet["B"+num].value, len(sheet["C"+num].value.split(","))])
        num = str(int(num)+1)
    return render(request, "Playlists.html", { 'playlists': playlists})
def playlists_create(request):
    global playlist_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
    if request.method == 'POST':
        songs1,songs2,songs3,songs4 = songs_create()
        songs = songs1 + songs2 + songs3 + songs4

        songs_backup = copy.deepcopy(songs)
        for i in range(len(songs_backup)-1,-1,-1):
            if request.POST.get("search", "").lower() not in songs_backup[i][0].lower():
                songs_backup.pop(i)
        songs1,songs2,songs3,songs4 = divide(songs_backup)
    else:
        songs1,songs2,songs3,songs4 = songs_create()
    songs = songs1 + songs2 + songs3 + songs4
    return render(request, "Playlist_Create.html", { 'all_songs': songs, 'playlist_songs': [] })
def playlists_view(request):
    global playlist
    global playing_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
    if request.method == 'POST':
        num=-1
        found = False
        while found == False:
            num = num + 1
            if "B"+str(num)+'B' in request.POST:
                found = True
        email =  request.POST.get("B"+str(num)+'B', "")
        name =  request.POST.get("B"+str(num)+'B1', "")
        workbook = openpyxl.load_workbook(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Playlists.xlsx")
        sheet = workbook.active
        num = 1
        while sheet["A" + str(num)].value!=email or sheet["B" + str(num)].value!=name:
            num = num + 1
        songs = sheet["C" + str(num)].value.split(",")
        for i in range(len(songs)):
            workbook2 = openpyxl.load_workbook(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Songs.xlsx")
            sheet2 = workbook2.active
            num = 1
            while sheet2["A" + str(num)].value!=songs[i]:
                num = num + 1
            songs[i] = [songs[i],sheet2["B" + str(num)].value,sheet2["C" + str(num)].value,sheet2["D" + str(num)].value,sheet2["E" + str(num)].value]
        playlist = [email,name,songs]
        playing_songs = copy.deepcopy(songs)
        songs1,songs2,songs3,songs4 = divide(playing_songs)
    return render(request, "Playlist_View.html", { 'name': name, 'songs1': songs1, 'songs2': songs2, 'songs3': songs3, 'songs4': songs4})
def add_song(request):
    global all_songs
    global playlist_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')

    if request.method == 'POST':
        num=-1
        found = False
        while found == False:
            num = num + 1
            if "B"+str(num)+'B' in request.POST:
                found = True

        song = [request.POST.get(str(num)+"song1", ""), request.POST.get(str(num)+"song2", ""), request.POST.get(str(num)+"song3", ""), request.POST.get(str(num)+"song4", "")]
        done = False
        for i in range(len(playlist_songs)):
            if playlist_songs[i][0] == song[0] and playlist_songs[i][1] == song[1]:
                done = True
                break
        if done == False:
            playlist_songs.append(song)
            for i in range(len(all_songs)):
                if all_songs[i][0] == song[0] and all_songs[i][1] == song[1]:
                    all_songs.pop(i)
                    break

    return render(request, "Playlist_Create.html", { 'all_songs': all_songs, 'playlist_songs': playlist_songs})
def remove_song(request):
    global all_songs
    global playlist_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')

    if request.method == 'POST':
        num=-1
        found = False
        while found == False:
            num = num + 1
            if "B"+str(num)+'B' in request.POST:
                found = True
        song = [request.POST.get(str(num)+"song1", ""), request.POST.get(str(num)+"song2", ""), request.POST.get(str(num)+"song3", ""), request.POST.get(str(num)+"song4", "")]
        
        done = False
        for i in range(len(all_songs)):
            if all_songs[i][0] == song[0] and all_songs[i][1] == song[1]:
                done = True
                break
        if done == False:
            all_songs.append(song)
            for i in range(len(playlist_songs)):
                if playlist_songs[i][0] == song[0] and playlist_songs[i][1] == song[1]:
                    playlist_songs.pop(i)
                    break

    return render(request, "Playlist_Create.html", { 'all_songs': all_songs, 'playlist_songs': playlist_songs})
def create(request):
    global all_songs
    global playlist_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')

    workbook = openpyxl.load_workbook(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Playlists.xlsx")
    sheet = workbook.active
    num = 1
    while sheet["A" + str(num)].value!=None:
        num = num + 1

    sheet["A"+str(num)] = user1[0]
    if request.method == 'POST':
        sheet["B"+str(num)] = request.POST.get("playlist_name", "")
    else:
        global name
        sheet["B"+str(num)] = name
    songs = ""
    for song in playlist_songs:
        songs = songs + song[0] + ","
    songs = songs[:len(songs)-1]
    sheet["C"+str(num)] = songs
    workbook.save(r"C:\Users\meyer\Desktop\Coding\Python\Website\OCRTunes\main\Playlists.xlsx")

    return HttpResponseRedirect('/main/playlists')
def playlists_genre(request):
    global name
    global playlist_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account') 
    songs1,songs2,songs3,songs4 = songs_create()
    playlist_songs = songs1 + songs2 + songs3 + songs4
    for i in range(len(playlist_songs)-1,-1,-1):
        if playlist_songs[i][5] != request.POST.get('genre', ''):
            playlist_songs.pop(i)
    name = request.POST.get('genre', '')
    if len(playlist_songs) == 0:
        return HttpResponseRedirect('/main/playlists')
    return HttpResponseRedirect('/main/create')
def playlists_length(request):
    global name
    global playlist_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account') 
    songs1,songs2,songs3,songs4 = songs_create()
    playlist_songs1 = songs1 + songs2 + songs3 + songs4

    playlist_songs = [playlist_songs1[1]]

    name = request.POST.get('length', '')


    return HttpResponseRedirect('/main/create')

def search(request):
    global playlist_songs
    global all_songs
    user1 = views2.user1
    if user1 == []:
        return HttpResponseRedirect('/account')
    if request.method == 'POST':

        songs_backup = copy.deepcopy(all_songs)
        for i in range(len(songs_backup)-1,-1,-1):
            
            if request.POST.get("search", "").lower() not in songs_backup[i][0].lower() and  request.POST.get("search", "").lower() not in songs_backup[i][1].lower():
                songs_backup.pop(i)
        songs1,songs2,songs3,songs4 = divide(songs_backup)
    else:
        songs1,songs2,songs3,songs4 = songs_create()
    songs = songs1 + songs2 + songs3 + songs4

    return render(request, "Playlist_Create.html", { 'all_songs': songs, 'playlist_songs': playlist_songs })

def random(request):
    return render(request, "Random.html")
